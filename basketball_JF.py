import sys
import csv
import json
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QGridLayout, QLabel, QPushButton,
                             QSpinBox, QLineEdit, QGroupBox, QMessageBox,
                             QTabWidget, QTableWidget, QTableWidgetItem,
                             QFileDialog, QComboBox, QHeaderView, QInputDialog,
                             QDialog, QDialogButtonBox, QFormLayout, QDateEdit,
                             QListWidget, QListWidgetItem, QSplitter, QFrame,
                             QProgressBar, QSizePolicy)
from PyQt5.QtCore import QTimer, Qt, QDate
from PyQt5.QtGui import QFont, QColor, QPalette, QIcon, QPixmap


class FoulDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择犯规类型")
        self.setModal(True)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                font-weight: bold;
                color: #333;
            }
            QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
        """)
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)

        form_layout = QFormLayout()
        self.foul_type = QComboBox()
        # 常见的篮球犯规类型
        foul_types = [
            "个人犯规", "技术犯规", "违反体育道德犯规",
            "取消比赛资格犯规", "双方犯规", "进攻犯规", "防守犯规"
        ]
        self.foul_type.addItems(foul_types)
        form_layout.addRow(QLabel("犯规类型:"), self.foul_type)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        # 美化按钮
        buttons.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                border: none;
                color: white;
                padding: 8px 16px;
                text-align: center;
                text-decoration: none;
                font-size: 14px;
                margin: 4px 2px;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
                color: #7f8c8d;
            }
        """)

        layout.addWidget(buttons)

    def get_foul_type(self):
        return self.foul_type.currentText()


class BasketballScoreboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.game_data = {
            'team_a': {'name': '队伍A', 'score': 0, 'fouls': 0, 'timeouts': 3, 'players': {}},
            'team_b': {'name': '队伍B', 'score': 0, 'fouls': 0, 'timeouts': 3, 'players': {}},
            'quarter': 1,
            'time_remaining': 12 * 60,  # 12分钟，以秒为单位
            'game_history': []
        }

        # 新增：赛程数据和球员数据
        self.schedule_data = []
        self.player_data = {}

        self.initUI()

    def initUI(self):
        # 设置窗口标题和大小
        self.setWindowTitle('篮球比赛积分计分程序')
        self.setGeometry(100, 50, 1600, 900)  # 增加窗口大小

        # 设置应用程序样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 8px;
                margin-top: 1ex;
                padding-top: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #333;
            }
            QLabel {
                color: #333;
            }
            QPushButton {
                background-color: #3498db;
                border: none;
                color: white;
                padding: 8px 16px;
                text-align: center;
                text-decoration: none;
                font-size: 14px;
                margin: 4px 2px;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
                color: #7f8c8d;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
            QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
            QTableWidget {
                gridline-color: #ccc;
                background-color: white;
                alternate-background-color: #f9f9f9;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 6px;
                border: 1px solid #2980b9;
                font-weight: bold;
            }
            QTabWidget::pane {
                border: 1px solid #cccccc;
                background: white;
            }
            QTabBar::tab {
                background: #e0e0e0;
                border: 1px solid #cccccc;
                padding: 8px 12px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom-color: white;
            }
        """)

        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局 - 使用QSplitter实现可调整大小的面板
        main_splitter = QSplitter(Qt.Horizontal)

        # 左侧比赛控制面板
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(5, 5, 5, 5)

        # 比赛信息显示
        game_info_group = QGroupBox("比赛信息")
        game_info_group.setStyleSheet("""
            QGroupBox {
                background-color: #e8f4f8;
            }
        """)
        game_info_layout = QHBoxLayout()

        self.quarter_label = QLabel("第1节")
        self.quarter_label.setFont(QFont("Arial", 16, QFont.Bold))
        self.quarter_label.setStyleSheet("color: #2c3e50;")

        quarter_control_layout = QHBoxLayout()
        prev_quarter_btn = QPushButton("◀ 上一节")
        next_quarter_btn = QPushButton("下一节 ▶")
        prev_quarter_btn.clicked.connect(self.prev_quarter)
        next_quarter_btn.clicked.connect(self.next_quarter)

        quarter_control_layout.addWidget(prev_quarter_btn)
        quarter_control_layout.addWidget(self.quarter_label)
        quarter_control_layout.addWidget(next_quarter_btn)

        game_info_layout.addLayout(quarter_control_layout)
        game_info_group.setLayout(game_info_layout)

        left_layout.addWidget(game_info_group)

        # 队伍信息区域 - 使用水平布局并排显示两队
        teams_widget = QWidget()
        teams_layout = QHBoxLayout(teams_widget)
        teams_layout.setContentsMargins(0, 0, 0, 0)

        # 队伍A信息
        team_a_group = QGroupBox("队伍A")
        team_a_group.setStyleSheet("""
            QGroupBox {
                background-color: #ffe6e6;
            }
        """)
        team_a_layout = QVBoxLayout()

        self.team_a_name = QLineEdit("队伍A")
        self.team_a_name.setFont(QFont("Arial", 12, QFont.Bold))
        self.team_a_name.textChanged.connect(lambda: self.update_team_name('A'))

        score_layout = QHBoxLayout()
        self.team_a_score = QLabel("0")
        self.team_a_score.setFont(QFont("Arial", 48, QFont.Bold))
        self.team_a_score.setAlignment(Qt.AlignCenter)
        self.team_a_score.setStyleSheet("color: #e74c3c;")

        self.team_a_fouls = QLabel("犯规: 0")
        self.team_a_fouls.setFont(QFont("Arial", 10))
        self.team_a_timeouts = QLabel("暂停: 3")
        self.team_a_timeouts.setFont(QFont("Arial", 10))

        score_layout.addWidget(self.team_a_fouls)
        score_layout.addStretch(1)
        score_layout.addWidget(self.team_a_score)
        score_layout.addStretch(1)
        score_layout.addWidget(self.team_a_timeouts)

        team_a_layout.addWidget(self.team_a_name)
        team_a_layout.addLayout(score_layout)

        # 队伍A得分按钮
        a_buttons_layout = QGridLayout()
        a_points_btn_1 = QPushButton("+1分")
        a_points_btn_2 = QPushButton("+2分")
        a_points_btn_3 = QPushButton("+3分")
        a_points_btn_1.clicked.connect(lambda: self.update_score('A', 1))
        a_points_btn_2.clicked.connect(lambda: self.update_score('A', 2))
        a_points_btn_3.clicked.connect(lambda: self.update_score('A', 3))

        a_buttons_layout.addWidget(a_points_btn_1, 0, 0)
        a_buttons_layout.addWidget(a_points_btn_2, 0, 1)
        a_buttons_layout.addWidget(a_points_btn_3, 0, 2)

        # 队伍A其他控制
        a_control_layout = QHBoxLayout()
        a_foul_btn = QPushButton("犯规")
        a_timeout_btn = QPushButton("暂停")
        a_foul_btn.clicked.connect(lambda: self.show_foul_dialog('A'))
        a_timeout_btn.clicked.connect(lambda: self.update_timeouts('A'))

        a_control_layout.addWidget(a_foul_btn)
        a_control_layout.addWidget(a_timeout_btn)

        team_a_layout.addLayout(a_buttons_layout)
        team_a_layout.addLayout(a_control_layout)
        team_a_group.setLayout(team_a_layout)

        # 队伍B信息
        team_b_group = QGroupBox("队伍B")
        team_b_group.setStyleSheet("""
            QGroupBox {
                background-color: #e6f7ff;
            }
        """)
        team_b_layout = QVBoxLayout()

        self.team_b_name = QLineEdit("队伍B")
        self.team_b_name.setFont(QFont("Arial", 12, QFont.Bold))
        self.team_b_name.textChanged.connect(lambda: self.update_team_name('B'))

        score_layout_b = QHBoxLayout()
        self.team_b_score = QLabel("0")
        self.team_b_score.setFont(QFont("Arial", 48, QFont.Bold))
        self.team_b_score.setAlignment(Qt.AlignCenter)
        self.team_b_score.setStyleSheet("color: #3498db;")

        self.team_b_fouls = QLabel("犯规: 0")
        self.team_b_fouls.setFont(QFont("Arial", 10))
        self.team_b_timeouts = QLabel("暂停: 3")
        self.team_b_timeouts.setFont(QFont("Arial", 10))

        score_layout_b.addWidget(self.team_b_fouls)
        score_layout_b.addStretch(1)
        score_layout_b.addWidget(self.team_b_score)
        score_layout_b.addStretch(1)
        score_layout_b.addWidget(self.team_b_timeouts)

        team_b_layout.addWidget(self.team_b_name)
        team_b_layout.addLayout(score_layout_b)

        # 队伍B得分按钮
        b_buttons_layout = QGridLayout()
        b_points_btn_1 = QPushButton("+1分")
        b_points_btn_2 = QPushButton("+2分")
        b_points_btn_3 = QPushButton("+3分")
        b_points_btn_1.clicked.connect(lambda: self.update_score('B', 1))
        b_points_btn_2.clicked.connect(lambda: self.update_score('B', 2))
        b_points_btn_3.clicked.connect(lambda: self.update_score('B', 3))

        b_buttons_layout.addWidget(b_points_btn_1, 0, 0)
        b_buttons_layout.addWidget(b_points_btn_2, 0, 1)
        b_buttons_layout.addWidget(b_points_btn_3, 0, 2)

        # 队伍B其他控制
        b_control_layout = QHBoxLayout()
        b_foul_btn = QPushButton("犯规")
        b_timeout_btn = QPushButton("暂停")
        b_foul_btn.clicked.connect(lambda: self.show_foul_dialog('B'))
        b_timeout_btn.clicked.connect(lambda: self.update_timeouts('B'))

        b_control_layout.addWidget(b_foul_btn)
        b_control_layout.addWidget(b_timeout_btn)

        team_b_layout.addLayout(b_buttons_layout)
        team_b_layout.addLayout(b_control_layout)
        team_b_group.setLayout(team_b_layout)

        # 将两队添加到水平布局
        teams_layout.addWidget(team_a_group)
        teams_layout.addWidget(team_b_group)

        left_layout.addWidget(teams_widget)

        # 右侧控制面板
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(5, 5, 5, 5)

        # 计时器部分
        timer_group = QGroupBox("比赛计时器")
        timer_group.setStyleSheet("""
            QGroupBox {
                background-color: #fff8e1;
            }
        """)
        timer_layout = QVBoxLayout()

        self.timer_display = QLabel("12:00")
        self.timer_display.setFont(QFont("Arial", 36, QFont.Bold))
        self.timer_display.setAlignment(Qt.AlignCenter)
        self.timer_display.setStyleSheet("color: #e67e22;")

        timer_control_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始")
        self.pause_btn = QPushButton("暂停")
        self.reset_btn = QPushButton("重置")

        self.start_btn.clicked.connect(self.start_timer)
        self.pause_btn.clicked.connect(self.pause_timer)
        self.reset_btn.clicked.connect(self.reset_timer)

        timer_control_layout.addWidget(self.start_btn)
        timer_control_layout.addWidget(self.pause_btn)
        timer_control_layout.addWidget(self.reset_btn)

        timer_layout.addWidget(self.timer_display)
        timer_layout.addLayout(timer_control_layout)
        timer_group.setLayout(timer_layout)

        # 球员管理
        player_group = QGroupBox("球员管理")
        player_layout = QVBoxLayout()

        player_select_layout = QHBoxLayout()
        self.team_select = QComboBox()
        self.team_select.addItems(["队伍A", "队伍B"])

        self.player_select = QComboBox()
        # 默认添加一些球员号码
        for i in range(1, 16):
            self.player_select.addItem(str(i))

        player_select_layout.addWidget(QLabel("队伍:"))
        player_select_layout.addWidget(self.team_select)
        player_select_layout.addWidget(QLabel("球员:"))
        player_select_layout.addWidget(self.player_select)

        player_stats_layout = QGridLayout()
        points_btn = QPushButton("得分")
        rebound_btn = QPushButton("篮板")
        assist_btn = QPushButton("助攻")
        steal_btn = QPushButton("抢断")
        block_btn = QPushButton("盖帽")
        turnover_btn = QPushButton("失误")
        foul_btn = QPushButton("犯规")

        points_btn.clicked.connect(lambda: self.add_player_stat('points'))
        rebound_btn.clicked.connect(lambda: self.add_player_stat('rebounds'))
        assist_btn.clicked.connect(lambda: self.add_player_stat('assists'))
        steal_btn.clicked.connect(lambda: self.add_player_stat('steals'))
        block_btn.clicked.connect(lambda: self.add_player_stat('blocks'))
        turnover_btn.clicked.connect(lambda: self.add_player_stat('turnovers'))
        foul_btn.clicked.connect(self.add_player_foul)

        player_stats_layout.addWidget(points_btn, 0, 0)
        player_stats_layout.addWidget(rebound_btn, 0, 1)
        player_stats_layout.addWidget(assist_btn, 0, 2)
        player_stats_layout.addWidget(steal_btn, 1, 0)
        player_stats_layout.addWidget(block_btn, 1, 1)
        player_stats_layout.addWidget(turnover_btn, 1, 2)
        player_stats_layout.addWidget(foul_btn, 2, 0, 1, 3)

        player_layout.addLayout(player_select_layout)
        player_layout.addLayout(player_stats_layout)
        player_group.setLayout(player_layout)

        # 比赛控制
        control_group = QGroupBox("比赛控制")
        control_layout = QVBoxLayout()

        reset_score_btn = QPushButton("重置比分")
        save_game_btn = QPushButton("保存比赛")
        load_game_btn = QPushButton("加载比赛")
        export_btn = QPushButton("导出数据")

        reset_score_btn.clicked.connect(self.reset_score)
        save_game_btn.clicked.connect(self.save_game)
        load_game_btn.clicked.connect(self.load_game)
        export_btn.clicked.connect(self.export_data)

        control_layout.addWidget(reset_score_btn)
        control_layout.addWidget(save_game_btn)
        control_layout.addWidget(load_game_btn)
        control_layout.addWidget(export_btn)
        control_group.setLayout(control_layout)

        right_layout.addWidget(timer_group)
        right_layout.addWidget(player_group)
        right_layout.addWidget(control_group)
        right_layout.addStretch(1)

        # 将左右面板添加到分割器
        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([1000, 400])  # 设置初始大小比例

        # 设置中央部件布局
        central_layout = QHBoxLayout(central_widget)
        central_layout.addWidget(main_splitter)

        # 初始化数据表格
        self.init_data_tabs()

        # 初始化计时器
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_timer)
        self.timer_running = False

    def init_data_tabs(self):
        # 创建数据标签页
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #cccccc;
                background: white;
            }
            QTabBar::tab {
                background: #e0e0e0;
                border: 1px solid #cccccc;
                padding: 8px 12px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom-color: white;
            }
        """)

        # 比赛记录表
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels(["时间", "节次", "队伍", "事件", "详情"])
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setAlternatingRowColors(True)

        # 球员数据表
        self.players_table = QTableWidget()
        self.players_table.setColumnCount(10)
        self.players_table.setHorizontalHeaderLabels(
            ["队伍", "号码", "得分", "篮板", "助攻", "抢断", "盖帽", "失误", "犯规", "犯规类型"])
        self.players_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.players_table.setAlternatingRowColors(True)

        # 赛程管理表
        self.schedule_tab = QWidget()
        schedule_layout = QVBoxLayout(self.schedule_tab)

        # 赛程导入按钮
        schedule_import_layout = QHBoxLayout()
        import_schedule_btn = QPushButton("导入赛程表")
        import_players_btn = QPushButton("导入球员信息")
        import_schedule_btn.clicked.connect(self.import_schedule)
        import_players_btn.clicked.connect(self.import_player_info)

        schedule_import_layout.addWidget(import_schedule_btn)
        schedule_import_layout.addWidget(import_players_btn)
        schedule_import_layout.addStretch(1)

        # 赛程表格
        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(5)
        self.schedule_table.setHorizontalHeaderLabels(["日期", "时间", "主场队伍", "客场队伍", "场地"])
        self.schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.schedule_table.doubleClicked.connect(self.start_game_from_schedule)
        self.schedule_table.setAlternatingRowColors(True)

        # 开始比赛按钮
        start_game_btn = QPushButton("开始选中比赛")
        start_game_btn.clicked.connect(self.start_game_from_schedule)

        schedule_layout.addLayout(schedule_import_layout)
        schedule_layout.addWidget(self.schedule_table)
        schedule_layout.addWidget(start_game_btn)

        self.tabs.addTab(self.history_table, "比赛记录")
        self.tabs.addTab(self.players_table, "球员数据")
        self.tabs.addTab(self.schedule_tab, "赛程管理")

        # 将标签页添加到主窗口（注意：不能重新设置centralWidget，否则会删除之前的UI对象）
        # 创建一个容器widget来放置标签页
        tab_container = QWidget()
        tab_layout = QHBoxLayout(tab_container)
        tab_layout.addWidget(self.tabs)
        
        # 将标签页容器添加到主分割器的左侧
        # 我们需要将标签页添加到主布局中，而不是重新设置centralWidget
        # 修改为将标签页添加到中央部件的布局中
        central_widget = self.centralWidget()
        central_layout = central_widget.layout()
        central_layout.addWidget(tab_container)

    def import_schedule(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "导入赛程表", "", "Excel文件 (*.xlsx *.xls)")
        if file_path:
            try:
                # 使用openpyxl读取Excel文件
                wb = load_workbook(filename=file_path)
                sheet = wb.active

                # 获取表头
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value)

                # 确保必要的列存在
                required_columns = ['日期', '时间', '主场队伍', '客场队伍']
                for col in required_columns:
                    if col not in headers:
                        QMessageBox.warning(self, "导入错误", f"Excel文件中缺少必要的列: {col}")
                        return

                # 提取数据
                self.schedule_data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # 跳过空行
                        game = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                game[header] = row[i]
                        self.schedule_data.append(game)

                # 更新赛程表格
                self.update_schedule_table()

                QMessageBox.information(self, "导入成功", "赛程表已成功导入！")
            except Exception as e:
                QMessageBox.critical(self, "导入失败", f"导入赛程表时出错：{str(e)}")

    def import_player_info(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "导入球员信息", "", "Excel文件 (*.xlsx *.xls)")
        if file_path:
            try:
                # 使用openpyxl读取Excel文件
                wb = load_workbook(filename=file_path)
                sheet = wb.active

                # 获取表头
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value)

                # 确保必要的列存在
                required_columns = ['队伍', '球衣号']
                for col in required_columns:
                    if col not in headers:
                        QMessageBox.warning(self, "导入错误", f"Excel文件中缺少必要的列: {col}")
                        return

                # 按队伍分组球员信息
                self.player_data = {}
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # 跳过空行
                        team_index = headers.index('队伍')
                        jersey_index = headers.index('球衣号')

                        team = row[team_index]
                        jersey = str(row[jersey_index]) if row[jersey_index] is not None else ""

                        if team and jersey:
                            if team not in self.player_data:
                                self.player_data[team] = []

                            if jersey not in self.player_data[team]:
                                self.player_data[team].append(jersey)

                QMessageBox.information(self, "导入成功", "球员信息已成功导入！")
            except Exception as e:
                QMessageBox.critical(self, "导入失败", f"导入球员信息时出错：{str(e)}")

    def update_schedule_table(self):
        self.schedule_table.setRowCount(len(self.schedule_data))

        for i, game in enumerate(self.schedule_data):
            self.schedule_table.setItem(i, 0, QTableWidgetItem(str(game.get('日期', ''))))
            self.schedule_table.setItem(i, 1, QTableWidgetItem(str(game.get('时间', ''))))
            self.schedule_table.setItem(i, 2, QTableWidgetItem(str(game.get('主场队伍', ''))))
            self.schedule_table.setItem(i, 3, QTableWidgetItem(str(game.get('客场队伍', ''))))
            self.schedule_table.setItem(i, 4, QTableWidgetItem(str(game.get('场地', ''))))

    def start_game_from_schedule(self):
        selected_row = self.schedule_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "选择错误", "请先选择一场比赛")
            return

        game = self.schedule_data[selected_row]
        home_team = game.get('主场队伍', '')
        away_team = game.get('客场队伍', '')

        if not home_team or not away_team:
            QMessageBox.warning(self, "数据错误", "赛程数据中缺少队伍信息")
            return

        # 设置队伍名称
        self.team_a_name.setText(home_team)
        self.team_b_name.setText(away_team)
        self.update_team_name('A')
        self.update_team_name('B')

        # 导入球员信息
        self.import_players_for_teams(home_team, away_team)

        # 重置比赛数据
        self.reset_score()

        # 切换到计分标签页
        self.tabs.setCurrentIndex(0)

        QMessageBox.information(self, "比赛准备就绪", f"已准备好比赛: {home_team} vs {away_team}")

    def import_players_for_teams(self, team_a, team_b):
        # 清空当前球员选择
        self.player_select.clear()

        # 导入队伍A的球员
        if team_a in self.player_data:
            for jersey in self.player_data[team_a]:
                self.player_select.addItem(jersey)

        # 导入队伍B的球员
        if team_b in self.player_data:
            for jersey in self.player_data[team_b]:
                if self.player_select.findText(jersey) == -1:  # 避免重复添加
                    self.player_select.addItem(jersey)

    def show_foul_dialog(self, team):
        dialog = FoulDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            foul_type = dialog.get_foul_type()
            self.update_fouls(team, foul_type)

    def update_team_name(self, team):
        if team == 'A':
            self.game_data['team_a']['name'] = self.team_a_name.text()
        else:
            self.game_data['team_b']['name'] = self.team_b_name.text()

    def update_score(self, team, points):
        if team == 'A':
            self.game_data['team_a']['score'] += points
            self.team_a_score.setText(str(self.game_data['team_a']['score']))
            team_name = self.game_data['team_a']['name']
        else:
            self.game_data['team_b']['score'] += points
            self.team_b_score.setText(str(self.game_data['team_b']['score']))
            team_name = self.game_data['team_b']['name']

        # 记录得分事件
        event = f"得分 +{points}"
        self.add_game_event(team_name, event)

    def update_fouls(self, team, foul_type="个人犯规"):
        if team == 'A':
            self.game_data['team_a']['fouls'] += 1
            self.team_a_fouls.setText(f"犯规: {self.game_data['team_a']['fouls']}")
            team_name = self.game_data['team_a']['name']
        else:
            self.game_data['team_b']['fouls'] += 1
            self.team_b_fouls.setText(f"犯规: {self.game_data['team_b']['fouls']}")
            team_name = self.game_data['team_b']['name']

        # 记录犯规事件
        self.add_game_event(team_name, f"犯规 ({foul_type})")

    def update_timeouts(self, team):
        if team == 'A':
            if self.game_data['team_a']['timeouts'] > 0:
                self.game_data['team_a']['timeouts'] -= 1
                self.team_a_timeouts.setText(f"暂停: {self.game_data['team_a']['timeouts']}")
                team_name = self.game_data['team_a']['name']
                # 记录暂停事件
                self.add_game_event(team_name, "暂停")
                # 暂停计时器
                self.pause_timer()
            else:
                QMessageBox.warning(self, "暂停次数不足", "队伍A没有剩余的暂停次数")
        else:
            if self.game_data['team_b']['timeouts'] > 0:
                self.game_data['team_b']['timeouts'] -= 1
                self.team_b_timeouts.setText(f"暂停: {self.game_data['team_b']['timeouts']}")
                team_name = self.game_data['team_b']['name']
                # 记录暂停事件
                self.add_game_event(team_name, "暂停")
                # 暂停计时器
                self.pause_timer()
            else:
                QMessageBox.warning(self, "暂停次数不足", "队伍B没有剩余的暂停次数")

    def add_player_stat(self, stat_type):
        team = self.team_select.currentText()
        player_number = self.player_select.currentText()

        if team == "队伍A":
            team_key = 'team_a'
        else:
            team_key = 'team_b'

        if player_number not in self.game_data[team_key]['players']:
            self.init_player_data(team_key, player_number)

        if stat_type == 'points':
            # 弹出对话框选择得分类型
            points_type, ok = QInputDialog.getItem(self, "得分类型", "选择得分类型:",
                                                   ["2分", "3分", "罚球"], 0, False)
            if ok:
                if points_type == "2分":
                    points = 2
                elif points_type == "3分":
                    points = 3
                else:
                    points = 1

                self.game_data[team_key]['players'][player_number]['points'] += points
                self.game_data[team_key]['score'] += points

                if team_key == 'team_a':
                    self.team_a_score.setText(str(self.game_data[team_key]['score']))
                else:
                    self.team_b_score.setText(str(self.game_data[team_key]['score']))

                # 记录得分事件
                event = f"{player_number}号球员得分 +{points}"
                self.add_game_event(self.game_data[team_key]['name'], event)
        else:
            self.game_data[team_key]['players'][player_number][stat_type] += 1

            # 记录其他统计事件
            stat_names = {
                'rebounds': '篮板', 'assists': '助攻', 'steals': '抢断',
                'blocks': '盖帽', 'turnovers': '失误'
            }
            event = f"{player_number}号球员{stat_names[stat_type]}"
            self.add_game_event(self.game_data[team_key]['name'], event)

        self.update_players_table()

    def add_player_foul(self):
        team = self.team_select.currentText()
        player_number = self.player_select.currentText()

        dialog = FoulDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            foul_type = dialog.get_foul_type()
            self.record_player_foul(team, player_number, foul_type)

    def record_player_foul(self, team, player_number, foul_type):
        if team == "队伍A":
            team_key = 'team_a'
        else:
            team_key = 'team_b'

        if player_number not in self.game_data[team_key]['players']:
            self.init_player_data(team_key, player_number)

        # 更新球员犯规数据
        if 'fouls' not in self.game_data[team_key]['players'][player_number]:
            self.game_data[team_key]['players'][player_number]['fouls'] = {}

        if foul_type not in self.game_data[team_key]['players'][player_number]['fouls']:
            self.game_data[team_key]['players'][player_number]['fouls'][foul_type] = 0

        self.game_data[team_key]['players'][player_number]['fouls'][foul_type] += 1

        # 更新队伍犯规计数
        self.game_data[team_key]['fouls'] += 1
        if team_key == 'team_a':
            self.team_a_fouls.setText(f"犯规: {self.game_data[team_key]['fouls']}")
        else:
            self.team_b_fouls.setText(f"犯规: {self.game_data['team_b']['fouls']}")

        # 记录犯规事件
        event = f"{player_number}号球员{foul_type}"
        self.add_game_event(self.game_data[team_key]['name'], event)

        self.update_players_table()

    def init_player_data(self, team_key, player_number):
        self.game_data[team_key]['players'][player_number] = {
            'points': 0, 'rebounds': 0, 'assists': 0,
            'steals': 0, 'blocks': 0, 'turnovers': 0,
            'fouls': {}
        }

    def add_game_event(self, team, event):
        current_time = self.timer_display.text()
        quarter = self.game_data['quarter']
        details = f"{team} - {event}"

        # 添加到历史记录
        self.game_data['game_history'].append({
            'time': current_time,
            'quarter': quarter,
            'team': team,
            'event': event,
            'details': details
        })

        # 更新历史表格
        self.update_history_table()

    def update_history_table(self):
        history = self.game_data['game_history']
        self.history_table.setRowCount(len(history))

        for i, event in enumerate(history):
            self.history_table.setItem(i, 0, QTableWidgetItem(event['time']))
            self.history_table.setItem(i, 1, QTableWidgetItem(f"第{event['quarter']}节"))
            self.history_table.setItem(i, 2, QTableWidgetItem(event['team']))
            self.history_table.setItem(i, 3, QTableWidgetItem(event['event']))
            self.history_table.setItem(i, 4, QTableWidgetItem(event['details']))

    def update_players_table(self):
        # 清空表格
        self.players_table.setRowCount(0)

        row = 0
        # 添加队伍A的球员数据
        for team_key in ['team_a', 'team_b']:
            team_name = self.game_data[team_key]['name']
            for player_number, stats in self.game_data[team_key]['players'].items():
                self.players_table.insertRow(row)
                self.players_table.setItem(row, 0, QTableWidgetItem(team_name))
                self.players_table.setItem(row, 1, QTableWidgetItem(player_number))
                self.players_table.setItem(row, 2, QTableWidgetItem(str(stats['points'])))
                self.players_table.setItem(row, 3, QTableWidgetItem(str(stats['rebounds'])))
                self.players_table.setItem(row, 4, QTableWidgetItem(str(stats['assists'])))
                self.players_table.setItem(row, 5, QTableWidgetItem(str(stats['steals'])))
                self.players_table.setItem(row, 6, QTableWidgetItem(str(stats['blocks'])))
                self.players_table.setItem(row, 7, QTableWidgetItem(str(stats['turnovers'])))

                # 计算总犯规数
                total_fouls = sum(stats.get('fouls', {}).values())
                self.players_table.setItem(row, 8, QTableWidgetItem(str(total_fouls)))

                # 显示犯规类型详情
                foul_details = ", ".join([f"{k}:{v}" for k, v in stats.get('fouls', {}).items()])
                self.players_table.setItem(row, 9, QTableWidgetItem(foul_details))

                row += 1

    def prev_quarter(self):
        if self.game_data['quarter'] > 1:
            self.game_data['quarter'] -= 1
            self.quarter_label.setText(f"第{self.game_data['quarter']}节")
            # 重置本节时间
            self.game_data['time_remaining'] = 12 * 60
            self.update_timer_display()

    def next_quarter(self):
        if self.game_data['quarter'] < 4:
            self.game_data['quarter'] += 1
        else:
            # 加时赛
            self.game_data['quarter'] += 1
            # 加时赛时间为5分钟
            self.game_data['time_remaining'] = 5 * 60

        self.quarter_label.setText(f"第{self.game_data['quarter']}节")
        self.update_timer_display()

    def start_timer(self):
        if not self.timer_running:
            self.timer.start(1000)  # 每秒更新一次
            self.timer_running = True

    def pause_timer(self):
        if self.timer_running:
            self.timer.stop()
            self.timer_running = False

    def reset_timer(self):
        self.pause_timer()
        if self.game_data['quarter'] <= 4:
            self.game_data['time_remaining'] = 12 * 60
        else:
            self.game_data['time_remaining'] = 5 * 60
        self.update_timer_display()

    def update_timer(self):
        if self.game_data['time_remaining'] > 0:
            self.game_data['time_remaining'] -= 1
            self.update_timer_display()
        else:
            self.pause_timer()
            QMessageBox.information(self, "时间到", f"第{self.game_data['quarter']}节结束！")

    def update_timer_display(self):
        minutes = self.game_data['time_remaining'] // 60
        seconds = self.game_data['time_remaining'] % 60
        self.timer_display.setText(f"{minutes:02d}:{seconds:02d}")

    def reset_score(self):
        reply = QMessageBox.question(self, "确认重置", "确定要重置比赛数据吗？所有数据将丢失！",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.game_data['team_a']['score'] = 0
            self.game_data['team_b']['score'] = 0
            self.game_data['team_a']['fouls'] = 0
            self.game_data['team_b']['fouls'] = 0
            self.game_data['team_a']['timeouts'] = 3
            self.game_data['team_b']['timeouts'] = 3
            self.game_data['team_a']['players'] = {}
            self.game_data['team_b']['players'] = {}
            self.game_data['quarter'] = 1
            self.game_data['time_remaining'] = 12 * 60
            self.game_data['game_history'] = []

            self.team_a_score.setText("0")
            self.team_b_score.setText("0")
            self.team_a_fouls.setText("犯规: 0")
            self.team_b_fouls.setText("犯规: 0")
            self.team_a_timeouts.setText("暂停: 3")
            self.team_b_timeouts.setText("暂停: 3")
            self.quarter_label.setText("第1节")
            self.update_timer_display()
            self.update_history_table()
            self.update_players_table()

    def save_game(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "保存比赛", "", "JSON文件 (*.json)")
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.game_data, f, ensure_ascii=False, indent=4)
                QMessageBox.information(self, "保存成功", "比赛数据已成功保存！")
            except Exception as e:
                QMessageBox.critical(self, "保存失败", f"保存比赛数据时出错：{str(e)}")

    def load_game(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "加载比赛", "", "JSON文件 (*.json)")
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.game_data = json.load(f)

                # 更新UI
                self.team_a_name.setText(self.game_data['team_a']['name'])
                self.team_b_name.setText(self.game_data['team_b']['name'])
                self.team_a_score.setText(str(self.game_data['team_a']['score']))
                self.team_b_score.setText(str(self.game_data['team_b']['score']))
                self.team_a_fouls.setText(f"犯规: {self.game_data['team_a']['fouls']}")
                self.team_b_fouls.setText(f"犯规: {self.game_data['team_b']['fouls']}")
                self.team_a_timeouts.setText(f"暂停: {self.game_data['team_a']['timeouts']}")
                self.team_b_timeouts.setText(f"暂停: {self.game_data['team_b']['timeouts']}")
                self.quarter_label.setText(f"第{self.game_data['quarter']}节")
                self.update_timer_display()
                self.update_history_table()
                self.update_players_table()

                QMessageBox.information(self, "加载成功", "比赛数据已成功加载！")
            except Exception as e:
                QMessageBox.critical(self, "加载失败", f"加载比赛数据时出错：{str(e)}")

    def export_data(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "导出数据", "", "CSV文件 (*.csv)")
        if file_path:
            try:
                # 导出比赛记录
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['时间', '节次', '队伍', '事件', '详情'])
                    for event in self.game_data['game_history']:
                        writer.writerow([
                            event['time'],
                            f"第{event['quarter']}节",
                            event['team'],
                            event['event'],
                            event['details']
                        ])

                # 导出球员数据
                player_file_path = file_path.replace('.csv', '_players.csv')
                with open(player_file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(
                        ['队伍', '号码', '得分', '篮板', '助攻', '抢断', '盖帽', '失误', '总犯规', '犯规详情'])

                    for team_key in ['team_a', 'team_b']:
                        team_name = self.game_data[team_key]['name']
                        for player_number, stats in self.game_data[team_key]['players'].items():
                            total_fouls = sum(stats.get('fouls', {}).values())
                            foul_details = "; ".join([f"{k}:{v}" for k, v in stats.get('fouls', {}).items()])

                            writer.writerow([
                                team_name,
                                player_number,
                                stats['points'],
                                stats['rebounds'],
                                stats['assists'],
                                stats['steals'],
                                stats['blocks'],
                                stats['turnovers'],
                                total_fouls,
                                foul_details
                            ])

                QMessageBox.information(self, "导出成功", "比赛数据已成功导出为CSV文件！")
            except Exception as e:
                QMessageBox.critical(self, "导出失败", f"导出比赛数据时出错：{str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)

    # 设置应用程序字体
    font = QFont("Microsoft YaHei", 9)  # 使用微软雅黑字体
    app.setFont(font)

    scoreboard = BasketballScoreboard()
    scoreboard.show()
    sys.exit(app.exec_())
